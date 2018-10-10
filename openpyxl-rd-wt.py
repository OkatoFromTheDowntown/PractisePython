import openpyxl

wb = openpyxl.Workbook()

print(wb.get_sheet_names())

dest_filename = r'target.xlsx'

sheet1 = wb.active
sheet1.title = 'new_sheet'
for row in range(4):
  sheet1.append(range(60))

sheet2 = wb.create_sheet(title='create_sheet')
sheet2['F5'] = 'Copr. Okato'

wb.save(dest_filename)

cpwb = openpyxl.load_workbook(dest_filename)
sheet = cpwb.get_sheet_by_name('create_sheet')
sheet.title = 'copyright_sheet'

cpwb.save('copy' + dest_filename)